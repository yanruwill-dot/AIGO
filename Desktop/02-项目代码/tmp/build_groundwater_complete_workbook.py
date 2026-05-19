from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


OUT_DIR = Path("/Users/will/Desktop/02-项目代码")
BASE = "全国产业园区地下水污染状况调查项目_完整版_全国数据资料"
XLSX = OUT_DIR / f"{BASE}.xlsx"
MD = OUT_DIR / f"{BASE}.md"

NAVY = "17365D"
BLUE = "1F4E79"
LIGHT_BLUE = "D9EAF7"
MID_BLUE = "9DC3E6"
GRAY = "F2F4F7"
TEXT = "1F2937"
GREEN = "D9EAD3"
AMBER = "FCE4D6"
RED = "F4CCCC"
WHITE = "FFFFFF"


sources = [
    {
        "来源ID": "S01",
        "标题": "关于政协十三届全国委员会第四次会议第4521号提案答复的函",
        "发布日期": "2021-12-02",
        "来源类型": "生态环境部/政策答复",
        "URL": "https://www.mee.gov.cn/xxgk2018/xxgk/xxgk13/202112/t20211202_962760.html",
        "可引用要点": "2021年3月生态环境部印发化工园区地下水环境状况调查评估工作方案；2020年底重点单位名录共1.3万余家。",
    },
    {
        "来源ID": "S02",
        "标题": "生态环境部地下水保护新闻发布/材料",
        "发布日期": "2019-11-30",
        "来源类型": "生态环境部/新闻材料",
        "URL": "https://www.mee.gov.cn/ywdt/hjnews/201911/t20191130_744924.shtml",
        "可引用要点": "2011-2017年开展全国地下水基础环境状况调查评估；2018-2019年开展典型工业园区周边地下水调查和风险评估。",
    },
    {
        "来源ID": "S03",
        "标题": "2021净土保卫战回顾",
        "发布日期": "2022-01-29",
        "来源类型": "地方生态环境厅转载/全国进展",
        "URL": "https://sthjj.cq.gov.cn/zwgk_249/gndt/202201/t20220129_10360258_wap.html",
        "可引用要点": "全国首批68个化工园区和9个重点铅锌矿区地下水调查评估外业调查于2021年12月完成。",
    },
    {
        "来源ID": "S04",
        "标题": "织密地下水污染防治网",
        "发布日期": "2025-04-02",
        "来源类型": "地方生态环境厅转载/全国进展",
        "URL": "https://sthjt.nx.gov.cn/xczx/gzdt/202504/t20250402_4871857.html",
        "可引用要点": "2021-2022年组织完成全国化工园区地下水环境状况调查评估；选择12个典型化工园区开展管控修复试点。",
    },
    {
        "来源ID": "S05",
        "标题": "生态环境部本级2026年部门预算",
        "发布日期": "2026-04-25",
        "来源类型": "生态环境部/部门预算PDF",
        "URL": "https://www.mee.gov.cn/ywgz/kjycw/bmyjsgl/202604/W020260425390266130785.pdf",
        "可引用要点": "2026年度绩效目标包括指导开展化工园区地下水污染专项整治数量399个，地下水国控点位I-IV类水比例不低于76%。",
    },
    {
        "来源ID": "S06",
        "标题": "云南省省级化工园区地下水基础环境状况调查评估",
        "发布日期": "2025-10-08",
        "来源类型": "项目承接单位案例",
        "URL": "https://www.yndky.com/document/detail/id/495.html",
        "可引用要点": "安宁、富民、澄江、寻甸、泸西5个园区；测绘1062.62km2；监测井163眼；采样检测953组；检测指标122项。",
    },
    {
        "来源ID": "S07",
        "标题": "湖北宜昌沿长江典型化工园区地下水调查评估项目通过验收",
        "发布日期": "2021-11-05",
        "来源类型": "地方地质局/项目验收",
        "URL": "https://dzj.hubei.gov.cn/qd/xwzx/dwxw/202111/t20211105_3845977.shtml",
        "可引用要点": "宜昌沿长江7个典型化工园区地下水环境状况调查评估项目通过验收，资金来源为中央水污染防治资金。",
    },
    {
        "来源ID": "S08",
        "标题": "云浮市完成第二批重点地下水污染源周边地下水基础环境状况调查评估",
        "发布日期": "2024-03-29",
        "来源类型": "地方生态环境局/项目进展",
        "URL": "https://www.yunfu.gov.cn/yfsthjj/gkmlpt/content/1/1801/post_1801997.html",
        "可引用要点": "调查10个重点地下水污染源，其中工业园区1个；调查面积约40km2；新布设永久性地下水监测井68个，其中工业园区16个。",
    },
    {
        "来源ID": "S09",
        "标题": "烟台市关于加强地下水入海监测提案答复",
        "发布日期": "2022-05-25",
        "来源类型": "地方生态环境局/提案答复",
        "URL": "https://www.yantai.gov.cn/art/2022/5/25/art_45948_2977230.html",
        "可引用要点": "2021年对6个化工园区、6个化工集聚区开展调查；现场踏勘97.65km2；水井调查164眼；筛选/新建监测井104眼；样品检测126件。",
    },
    {
        "来源ID": "S10",
        "标题": "辽宁省化工园区地下水环境状况调查评估工作全面启动",
        "发布日期": "2021-11-01",
        "来源类型": "省生态环境厅/项目启动",
        "URL": "https://sthj.ln.gov.cn/sthj/zwdt/snyw/E21EBDECD5724ABDA7A7AC43CD293631/index.shtml",
        "可引用要点": "2021年10月28日营口仙人岛化工园区第一口监测井正式开钻；营口、沈阳开始采样工作。",
    },
    {
        "来源ID": "S11",
        "标题": "辽宁省生态环境厅2022年度部门决算",
        "发布日期": "2023-09-01",
        "来源类型": "省财政/部门决算PDF",
        "URL": "https://www.ln.gov.cn/web/zwgkx/sjjhczbg/czjs/bmjs/202287/zfbm/2023090109265378604/2023090112073326087.pdf",
        "可引用要点": "化工园区地下水调查项目全年预算66万元、执行43万元；目标32个化工园区，完成21个。",
    },
    {
        "来源ID": "S12",
        "标题": "辽阳芳烃及精细化工产业化基地地下水污染详细调查项目审批",
        "发布日期": "2023-06-15",
        "来源类型": "省生态环境厅/项目审批",
        "URL": "https://sthj.ln.gov.cn/sthj/xxgk/zwxxgk/xxgkml/2023061509461249487/index.shtml",
        "可引用要点": "开展地下水污染详细调查和污染成因分析，明确污染羽、污染成因和风险管控/修复方案。",
    },
    {
        "来源ID": "S13",
        "标题": "河南省强化沿黄工业园区水污染治理",
        "发布日期": "2023-11-20",
        "来源类型": "省生态环境厅/工作进展",
        "URL": "https://sthjt.henan.gov.cn/2023/11-20/2850423.html",
        "可引用要点": "完成37个国家、省级化工园区地下水环境调查评估及成果集成，推进分级管理。",
    },
    {
        "来源ID": "S14",
        "标题": "2021年陕西省生态环境状况公报",
        "发布日期": "2022-06-02",
        "来源类型": "省生态环境厅/环境状况公报",
        "URL": "https://sthjt.shaanxi.gov.cn/xxgk/fdnr/hjzl/hjgb/202206/t20220602_2982011.html",
        "可引用要点": "陕西省启动化工园区地下水环境状况调查评估，完成4个化工园区调查评估。",
    },
    {
        "来源ID": "S15",
        "标题": "陕西省生态环境厅新闻发布材料",
        "发布日期": "2025-12-16",
        "来源类型": "省生态环境厅/新闻发布",
        "URL": "https://sthjt.shaanxi.gov.cn/xxgk/xwfb/202512/t20251216_3596535.html",
        "可引用要点": "陕西完成23个化工产业集聚区地下水污染状况调查评估，推动成果应用。",
    },
    {
        "来源ID": "S16",
        "标题": "六安市叶集化工园区地下水环境状况调查评估项目专家评审",
        "发布日期": "未披露",
        "来源类型": "地方生态环境局/项目评审",
        "URL": "https://sthjj.luan.gov.cn/hbyw/stbh/5041391.html",
        "可引用要点": "叶集化工园区调查方案和成果报告通过专家评审；工作包括建井、采样、成果报告。",
    },
    {
        "来源ID": "S17",
        "标题": "上海青浦区地下水环境状况调查评估工作推进",
        "发布日期": "2022-07-27",
        "来源类型": "区生态环境局/工作推进",
        "URL": "https://www.shqp.gov.cn/env/stzwgk/ml/yw/20220727/955438.html",
        "可引用要点": "建立一园一档、一企一档、一园一策、一园一报告，收集企业和监测井资料。",
    },
    {
        "来源ID": "S18",
        "标题": "莆田石门澳化工新材料产业园地下水调查评估采购",
        "发布日期": "2022-04-26",
        "来源类型": "公共资源交易/采购文件",
        "URL": "https://ggzyjy.xzfwzx.putian.gov.cn/ptsq/005002/005002003/005002003001/20220426/651f4ec3-4c27-479d-b405-6aaf26ad4d1a.html",
        "可引用要点": "预算165万元；园区及周边1km；不少于36个钻探采样孔；土壤样品108+11平行；地下水样品39+4平行；监测井39个。",
    },
    {
        "来源ID": "S19",
        "标题": "临湘高新区地下水污染状况详细调查项目公告",
        "发布日期": "2025-04-29",
        "来源类型": "地方政府/项目信息",
        "URL": "https://www.linxiang.gov.cn/24733/24760/24821/24902/26931/content_2327083.html",
        "可引用要点": "针对滨江片区、三湾片区开展地下水污染详细调查，查明污染物种类、浓度和空间分布并开展风险评估。",
    },
    {
        "来源ID": "S20",
        "标题": "湖南省中央生态环境资金项目储备库拟推荐项目公示",
        "发布日期": "2025-03-05",
        "来源类型": "省生态环境厅/资金储备",
        "URL": "https://sthjt.hunan.gov.cn/xxgk/tzgg/gg/202503/t20250305_33603003.html",
        "可引用要点": "临湘高新技术产业开发区地下水污染状况详细调查项目拟纳入中央储备库。",
    },
    {
        "来源ID": "S21",
        "标题": "苏尼特左旗工业园区地下水环境状况调查评估招标文件",
        "发布日期": "2024-04-22",
        "来源类型": "政府采购/招标PDF",
        "URL": "https://www.ccgp-neimenggu.gov.cn/gpx-bid-file/152523/gpx-tender/2024/4/22/402881cc8ec1d157018f0522be7f43ef.pdf?accessCode=9de828e2e1c7922f89dff1d1d5ef3a0c",
        "可引用要点": "工作内容包括资料收集、遥感解译、现场调查、水文地质钻探/建井、样品检测、数值模拟等。",
    },
    {
        "来源ID": "S22",
        "标题": "桐乡经济开发区化工园区土壤和地下水调查评估采购",
        "发布日期": "2023-06-01",
        "来源类型": "政府采购/采购公告",
        "URL": "https://jxszwsjb.jiaxing.gov.cn/art/2023/6/1/art_1229744161_377926.html",
        "可引用要点": "预算150万元；调查结果录入全国地下水基础环境状况调查评估子系统。",
    },
    {
        "来源ID": "S23",
        "标题": "淮南经济技术开发区化工园区地下水环境状况详细调查项目",
        "发布日期": "未披露",
        "来源类型": "地方政府/采购公告",
        "URL": "https://jkq.huainan.gov.cn/zwgk/zbcg/551762571.html",
        "可引用要点": "预算28万元；针对2021年省级调查发现的超标点位开展详细调查评估。",
    },
    {
        "来源ID": "S24",
        "标题": "苏州市2022年化工园区地下水环境状况调查评估质控项目",
        "发布日期": "2023-03",
        "来源类型": "政府采购/合同公告",
        "URL": "https://www.sdffwdwz.com/szhbj/szh/202303/3478a0dab7534e17834e8c0583e7a595.shtml",
        "可引用要点": "合同金额74.88万元；对资料收集、方案编制、建井采样、样品检测、报告编制等进行质控/抽查。",
    },
]


project_rows = [
    ["全国", "全国", "化工园区地下水环境状况调查评估工作方案", "政策部署", "2021-03", "全国化工园区", "发布工作方案", None, None, None, None, "形成全国统一调查评估口径", "已发布", "S01", "高"],
    ["全国", "全国", "全国首批化工园区地下水调查评估", "外业调查", "2021-12", "首批化工园区/重点铅锌矿区", "68个化工园区；9个重点铅锌矿区", None, None, None, None, "首批外业调查完成", "已完成", "S03", "中"],
    ["全国", "全国", "全国化工园区地下水环境状况调查评估", "调查评估", "2021-2022", "全国化工园区", "全国范围完成", None, None, None, None, "部分化工园区污染问题突出，进入管控修复试点", "已完成", "S04", "中"],
    ["全国", "全国", "化工园区地下水污染专项整治", "专项整治", "2026", "化工园区", "399个", None, None, None, None, "地下水国控点位I-IV类比例目标≥76%", "年度绩效目标", "S05", "高"],
    ["云南", "昆明/玉溪/红河", "云南省省级化工园区地下水基础环境状况调查评估", "调查评估", "2021-2025披露", "安宁、富民、澄江、寻甸、泸西5个省级园区", "5个园区", 1062.62, 163, 953, 122, "厘清水文地质单元、污染源、特征污染物并完成质量/污染评价", "成果披露", "S06", "中"],
    ["湖北", "宜昌", "沿长江7个典型化工园区地下水环境状况调查评估", "调查评估/验收", "2021-11", "宜昌沿长江7个典型化工园区", "7个园区", None, None, None, None, "项目通过验收，支撑沿江化工园区监管", "已验收", "S07", "高"],
    ["广东", "云浮", "第二批重点地下水污染源周边地下水基础环境状况调查评估", "调查评估", "2023-07至2024-03", "10个重点地下水污染源，其中工业园区1个", "工业园区1个；新井68个，其中工业园区16个", 40, 16, None, None, "形成重点源周边永久性监测井网络", "已完成", "S08", "高"],
    ["山东", "烟台", "烟台市化工园区/集聚区地下水环境状况调查", "调查评估", "2021", "6个化工园区、6个化工集聚区", "12个对象；现场踏勘97.65km2；水井调查164眼；重点区域2.3252km2", 97.65, 104, 126, None, "初步了解化工园区及周边水质现状，裕龙岛取得本底数据", "已完成", "S09", "高"],
    ["辽宁", "营口/沈阳等", "辽宁省化工园区地下水环境状况调查评估", "启动/调查", "2021-10", "营口仙人岛等化工园区", "监测井开钻；营口、沈阳开始采样", None, None, None, None, "建立园区地下水环境监测网络", "已启动", "S10", "高"],
    ["辽宁", "全省", "辽宁省化工园区地下水调查项目", "省级项目", "2022", "辽宁化工园区", "目标32个，完成21个", None, None, None, None, "受疫情和冬季冻土影响未全部完成", "部分完成", "S11", "高"],
    ["辽宁", "辽阳", "辽阳芳烃及精细化工基地地下水污染详细调查", "详细调查/溯源", "2023", "辽阳芳烃及精细化工产业化基地", "未披露", None, None, None, None, "明确污染羽、成因和风险管控/修复方案", "审批通过", "S12", "高"],
    ["河南", "全省/沿黄", "河南国家、省级化工园区地下水调查评估及成果集成", "成果集成", "2023披露", "国家、省级化工园区", "37个园区", None, None, None, None, "推进化工园区地下水污染防治分级管理", "已完成", "S13", "高"],
    ["陕西", "全省", "陕西省化工园区地下水环境状况调查评估", "调查评估", "2021", "化工园区", "完成4个园区", None, None, None, None, "启动省级化工园区调查评估", "已完成", "S14", "高"],
    ["陕西", "全省", "陕西化工产业集聚区地下水污染状况调查评估", "调查评估/成果应用", "2025披露", "化工产业集聚区", "23个集聚区", None, None, None, None, "推动调查评估成果应用", "已完成", "S15", "中"],
    ["安徽", "六安叶集", "叶集化工园区地下水环境状况调查评估", "方案/成果评审", "未披露", "叶集化工园区", "未披露", None, None, None, None, "方案与成果报告专家评审，覆盖建井和采样环节", "通过评审", "S16", "高"],
    ["上海", "青浦", "青浦区地下水环境状况调查评估", "资料建档/方案", "2022-07", "园区和企业", "未披露", None, None, None, None, "建立一园一档、一企一档、一园一策、一园一报告", "推进中", "S17", "高"],
    ["福建", "莆田", "石门澳化工新材料产业园地下水环境状况调查评估", "采购/调查", "2022-04", "园区及周边1km", "不少于36个钻探采样孔；39个监测井", None, 39, 162, 57, "采购文件披露详细工作量和检测指标组", "采购阶段", "S18", "高"],
    ["湖南", "岳阳临湘", "临湘高新区地下水污染状况详细调查", "详细调查", "2025", "滨江片区、三湾片区", "未披露", None, None, None, None, "查明污染物种类、浓度、空间分布并开展风险评估", "拟入库/公告", "S19/S20", "高"],
    ["内蒙古", "锡林郭勒苏尼特左旗", "苏尼特左旗工业园区地下水环境状况调查评估", "采购/调查", "2024-04", "苏尼特左旗工业园区", "未披露", None, None, None, None, "工作含遥感解译、钻探建井、采样检测和数值模拟", "招标阶段", "S21", "高"],
    ["浙江", "嘉兴桐乡", "桐乡经济开发区化工园区土壤和地下水调查评估", "采购/调查", "2023-06", "桐乡经济开发区化工园区", "未披露", None, None, None, None, "成果录入全国地下水基础环境状况调查评估子系统", "采购阶段", "S22", "高"],
    ["安徽", "淮南", "淮南经开区化工园区地下水环境状况详细调查", "详细调查", "未披露", "2021省级调查发现超标点位", "未披露", None, None, None, None, "针对超标点位开展详细调查评估", "采购阶段", "S23", "中"],
    ["江苏", "苏州", "苏州市2022年化工园区地下水调查评估质控", "质量控制", "2023-03", "苏州市化工园区调查评估工作", "未披露", None, None, None, None, "对资料、方案、建井采样、检测、报告进行质控/抽查", "合同阶段", "S24", "中"],
]


procurement_rows = [
    ["生态环境部本级土壤与农业农村环境质量监督管理", 80.42, "土壤污染源头防控、美丽乡村、地下水管理", "指导化工园区地下水污染专项整治399个", "未披露", "未披露", "地下水国控点位I-IV类≥76%", "年度绩效目标", "S05"],
    ["辽宁省化工园区地下水调查项目", 66.00, "全省化工园区", "目标32个园区，完成21个", "未披露", "未披露", "未披露", "执行43万元", "S11"],
    ["莆田石门澳化工新材料产业园调查评估", 165.00, "园区及周边1km", "不少于36个钻探采样孔；39个监测井", "土壤108件+平行11件", "地下水39件+平行4件", "地下水常规35项+离子/CO2 6项+特征7项+现场9项", "采购文件披露", "S18"],
    ["桐乡经济开发区化工园区调查评估", 150.00, "化工园区土壤和地下水", "未披露", "未披露", "未披露", "未披露", "成果录入全国子系统", "S22"],
    ["淮南经开区详细调查评估", 28.00, "2021省级调查发现超标点位", "未披露", "未披露", "未披露", "未披露", "详细调查评估", "S23"],
    ["苏州市2022年调查评估质控", 74.88, "苏州化工园区调查评估质控", "抽查建井采样", "抽查", "抽查", "抽查检测过程", "合同金额74.88万元", "S24"],
]


province_names = [
    "北京", "天津", "河北", "山西", "内蒙古", "辽宁", "吉林", "黑龙江", "上海", "江苏", "浙江", "安徽", "福建", "江西", "山东", "河南", "湖北", "湖南", "广东", "广西", "海南", "重庆", "四川", "贵州", "云南", "西藏", "陕西", "甘肃", "青海", "宁夏", "新疆"
]


province_overrides = {
    "全国": ["全国", "政策+进展+专项整治", 399, None, None, None, None, "2026年专项整治399个；2021-2022完成全国调查评估", "S01/S03/S04/S05", "高", "补充分省整治清单和污染等级分布"],
    "云南": ["云南", "省级项目案例", 5, 1062.62, 163, 953, None, "5个省级产业/化工园区工作量披露最完整", "S06", "中", "补充政府验收或采购原始文件"],
    "湖北": ["湖北", "地方验收", 7, None, None, None, None, "宜昌沿长江7个典型化工园区通过验收", "S07", "高", "补充各园区井数/样品/污染结论"],
    "广东": ["广东", "重点源调查", 1, 40, 16, None, None, "云浮1个工业园区纳入双源调查，新井16个", "S08", "高", "补充工业园区名称和检测指标"],
    "山东": ["山东", "市级调查", 12, 97.65, 104, 126, None, "烟台6个化工园区+6个化工集聚区", "S09", "高", "补充污染因子和评价等级"],
    "辽宁": ["辽宁", "省级项目+地方详细调查", 32, None, None, None, 66, "目标32个、完成21个；预算66万元、执行43万元", "S10/S11/S12", "高", "补充21个已完成园区清单"],
    "河南": ["河南", "省级成果集成", 37, None, None, None, None, "完成37个国家/省级化工园区调查评估及成果集成", "S13", "高", "补充分级管理结果"],
    "陕西": ["陕西", "省级进展", 23, None, None, None, None, "2021完成4个；后续披露23个化工产业集聚区", "S14/S15", "中", "核验2025发布材料并补充清单"],
    "安徽": ["安徽", "项目评审/详细调查", 2, None, None, None, 28, "叶集评审、淮南超标点位详细调查", "S16/S23", "中", "补充安徽省级总量口径"],
    "上海": ["上海", "区级资料建档", None, None, None, None, None, "青浦推进一园一档、一企一档、一园一策、一园一报告", "S17", "高", "补充园区数量和井点数据"],
    "福建": ["福建", "采购工作量完整", 1, None, 39, 162, 165, "莆田石门澳披露钻孔、监测井、土壤/地下水样品和指标", "S18", "高", "补充验收结果"],
    "湖南": ["湖南", "详细调查项目", 1, None, None, None, None, "临湘高新区详细调查拟纳入中央储备库", "S19/S20", "高", "补充预算、井数、样品"],
    "内蒙古": ["内蒙古", "招标文件", 1, None, None, None, None, "苏尼特左旗工业园区招标披露工作模块", "S21", "高", "补充合同额和工作量"],
    "浙江": ["浙江", "采购公告", 1, None, None, None, 150, "桐乡经开区调查评估预算150万元", "S22", "高", "补充井数和样品数"],
    "江苏": ["江苏", "质控合同", None, None, None, None, 74.88, "苏州调查评估质控合同74.88万元", "S24", "中", "补充苏州市园区总数和项目成果"],
}


def num_or_none(v):
    return v if isinstance(v, (int, float)) else None


province_rows = [province_overrides["全国"]]
for name in province_names:
    if name in province_overrides:
        province_rows.append(province_overrides[name])
    else:
        province_rows.append([name, "本轮未检索到公开量化数据", None, None, None, None, None, "本轮公开来源未形成可核验数字；不得外推填数", "未匹配", "低", "继续检索省生态环境厅、政府采购、环境公报"])


method_rows = [
    ["一园一档", "园区基础信息、企业清单、污染源、监测井、历史调查成果归档", "前期资料收集/监管建档", "上海青浦、云南案例", "S17/S06"],
    ["一企一档", "园区内重点企业原辅材料、排污、危废、历史泄漏和地下设施资料", "污染源识别", "上海青浦", "S17"],
    ["水文地质单元", "厘清边界、含水层结构、补径排条件和地下水流场", "调查评价基础", "云南案例", "S06"],
    ["特征污染物识别", "结合行业、原辅材料、排污历史和检测结果识别", "监测指标优化/风险判别", "云南/临湘", "S06/S19"],
    ["监测井建设", "按上游/源区/下游/边界等布设永久或临时井", "长期监测网络", "云南/云浮/莆田", "S06/S08/S18"],
    ["采样检测", "地下水、土壤、地表水等样品采集、保存、流转、检测", "污染评价依据", "云南/烟台/莆田", "S06/S09/S18"],
    ["质控抽查", "对方案、建井、采样、检测、报告进行外部复核", "成果可信度控制", "苏州质控项目", "S24"],
    ["污染羽/成因分析", "通过详细调查和模拟确定污染范围、迁移路径、成因", "风险管控/修复方案", "辽阳/临湘", "S12/S19"],
    ["成果入库", "录入全国地下水基础环境状况调查评估子系统", "全国汇总和监管", "桐乡采购", "S22"],
]


risk_rows = [
    ["源头识别不清", "园区企业多、历史资料缺、地下管线复杂", "一园一档/一企一档+历史影像+企业访谈", "资料完整率、重点企业覆盖率", "S06/S17"],
    ["污染羽边界不清", "初查井点不足或水文地质单元复杂", "加密监测井+水文地质试验+数值模拟", "边界井达标率、污染羽闭合度", "S12/S21"],
    ["监测网络不可持续", "临时调查井不能支撑长期预警", "永久性监测井网络+分级监测频次", "永久井数量、有效数据率", "S08/S05"],
    ["成果难以进入治理", "调查报告停留在评价层，缺少工程化路径", "风险分级+一园一策+中央资金储备", "进入储备库项目数、整治完成率", "S04/S20"],
    ["数据质量不稳定", "多单位施工检测导致口径不一", "第三方质控/平行样/现场抽查/报告复核", "平行样合格率、质控问题闭环率", "S18/S24"],
]


field_rows = [
    ["字段", "说明", "填报规则"],
    ["省份", "项目所在省级行政区或全国口径", "全国政策/预算用“全国”"],
    ["城市/地区", "城市、区县、园区所在地区", "公开来源未披露则填“未披露”"],
    ["项目名称", "公开来源中的项目或工作名称", "尽量保留原始口径"],
    ["阶段", "政策部署、调查评估、详细调查、质控、验收、专项整治等", "按来源语义归类"],
    ["对象范围", "园区、集聚区、重点源、周边范围", "不从相似项目外推"],
    ["调查面积(km²)", "公开披露调查/踏勘/测绘面积", "只填可核验数字"],
    ["监测井数量", "新建、筛选或永久监测井数量", "若来源区分多个口径，在备注中说明"],
    ["采样数量", "水、土、地表水等样品总量或组合量", "组合量需在证据摘要解释"],
    ["检测指标", "检测项目数或指标组", "指标组无法折算时可用文字"],
    ["可信度", "高/中/低", "政府/采购源通常高；承接单位案例中；未匹配低"],
]


timeline_rows = [
    ["2011-2017", "全国地下水基础环境状况调查评估", "全国地下水基础环境底账形成", "S02"],
    ["2018-2019", "典型工业园区周边地下水调查和风险评估", "工业园区成为重点风险对象", "S02"],
    ["2021-03", "生态环境部印发化工园区地下水调查评估工作方案", "全国统一技术和工作口径", "S01"],
    ["2021-12", "首批68个化工园区外业调查完成", "调查评估进入批量推进阶段", "S03"],
    ["2021-2022", "全国化工园区调查评估完成", "由摸底进入问题识别和管控修复", "S04"],
    ["2023-2025", "省市详细调查、质控和资金储备项目增多", "从初查转向详细调查、溯源和工程储备", "S12/S18/S19/S20/S24"],
    ["2026", "399个化工园区专项整治列入绩效目标", "治理主线从调查评估进入专项整治", "S05"],
]


def add_sheet_table(ws, rows, table_name):
    if not rows:
        return
    for row in rows:
        ws.append(row)
    end_col = get_column_letter(ws.max_column)
    ref = f"A1:{end_col}{ws.max_row}"
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ref


def style_sheet(ws, widths=None):
    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color=WHITE, bold=True, size=10)
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.font = Font(color=TEXT, size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    else:
        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 45
    ws.sheet_view.showGridLines = False


def link_sources(ws, source_col_name="来源ID"):
    source_map = {s["来源ID"]: s["URL"] for s in sources}
    headers = [c.value for c in ws[1]]
    if source_col_name not in headers:
        return
    col = headers.index(source_col_name) + 1
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row, col).value
        if not value:
            continue
        first = str(value).split("/")[0]
        url = source_map.get(first)
        if url:
            ws.cell(row, col).hyperlink = url
            ws.cell(row, col).style = "Hyperlink"


wb = Workbook()
wb.remove(wb.active)

dash = wb.create_sheet("管理看板")
dash.sheet_view.showGridLines = False
dash["A1"] = "全国产业园区/化工园区地下水污染状况调查项目数据表 - 完整版"
dash["A1"].font = Font(bold=True, size=18, color=WHITE)
dash["A1"].fill = PatternFill("solid", fgColor=NAVY)
dash.merge_cells("A1:H1")
dash["A3"] = "结论先行"
dash["A3"].font = Font(bold=True, size=13, color=NAVY)
insights = [
    "1. 全国主线已经从“调查评估”切换到“专项整治”：2026年绩效目标明确指导399个化工园区地下水污染专项整治。",
    "2. 已公开量化数据呈现强烈分化：云南、烟台、莆田等披露面积/井/样品/指标较完整，多数省份只披露项目阶段或未披露。",
    "3. 项目能力栈从普通采样扩展到一园一档、水文地质建模、污染羽识别、质控抽查和治理资金储备。",
    "4. 下一步最有价值的补数方向是：分省园区清单、井点/样品/污染因子、风险等级、专项整治完成率和治理投资。",
]
for idx, text in enumerate(insights, 4):
    dash.cell(idx, 1, text)
    dash.cell(idx, 1).alignment = Alignment(wrap_text=True, vertical="top")
    dash.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=8)

kpis = [
    ["KPI", "数值", "口径", "来源"],
    ["2026专项整治", 399, "化工园区地下水污染专项整治数量", "S05"],
    ["国控点位目标", "≥76%", "地下水国控点位I-IV类比例", "S05"],
    ["云南监测井", 163, "5个省级产业/化工园区", "S06"],
    ["云南采样检测", 953, "水、土样品采集及检测", "S06"],
    ["烟台对象", 12, "6个化工园区+6个化工集聚区", "S09"],
    ["河南成果集成", 37, "国家、省级化工园区", "S13"],
    ["福建莆田预算", 165, "万元，石门澳园区调查评估", "S18"],
]
for r, row in enumerate(kpis, 10):
    for c, val in enumerate(row, 1):
        dash.cell(r, c, val)
for cell in dash[10]:
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.font = Font(bold=True, color=WHITE)
for row in dash.iter_rows(min_row=11, max_row=17, min_col=1, max_col=4):
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=GRAY)
        cell.border = Border(left=Side(style="thin", color="D9E2F3"), right=Side(style="thin", color="D9E2F3"), top=Side(style="thin", color="D9E2F3"), bottom=Side(style="thin", color="D9E2F3"))
        cell.alignment = Alignment(wrap_text=True, vertical="top")
for col, width in {"A": 22, "B": 16, "C": 48, "D": 16, "E": 2, "F": 18, "G": 18, "H": 18}.items():
    dash.column_dimensions[col].width = width
dash.freeze_panes = "A10"

charts_data_start = 20
dash.cell(charts_data_start, 1, "省份/项目")
dash.cell(charts_data_start, 2, "园区/对象数")
chart_rows = [
    ["全国专项整治", 399],
    ["河南", 37],
    ["辽宁目标", 32],
    ["陕西披露", 23],
    ["烟台", 12],
    ["湖北宜昌", 7],
    ["云南", 5],
]
for idx, row in enumerate(chart_rows, charts_data_start + 1):
    dash.cell(idx, 1, row[0])
    dash.cell(idx, 2, row[1])
chart = BarChart()
chart.title = "公开披露园区/对象数量对比"
chart.y_axis.title = "数量"
chart.x_axis.title = "地区/项目"
data = Reference(dash, min_col=2, min_row=charts_data_start, max_row=charts_data_start + len(chart_rows))
cats = Reference(dash, min_col=1, min_row=charts_data_start + 1, max_row=charts_data_start + len(chart_rows))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 7
chart.width = 14
chart.style = 10
dash.add_chart(chart, "D20")
link_sources(dash, "来源")

timeline = wb.create_sheet("全国主线")
add_sheet_table(timeline, [["时间", "全国推进节点", "管理含义", "来源ID"]] + timeline_rows, "tbl_timeline")
style_sheet(timeline, {"A": 16, "B": 42, "C": 52, "D": 16})
link_sources(timeline)

province = wb.create_sheet("省份覆盖清单")
add_sheet_table(
    province,
    [["省份", "公开阶段", "园区/对象数", "调查面积(km²)", "监测井数量", "采样数量", "预算/合同额(万元)", "公开结论/备注", "来源ID", "可信度", "待补数据"]] + province_rows,
    "tbl_province_coverage",
)
style_sheet(province, {"A": 12, "B": 22, "C": 14, "D": 16, "E": 14, "F": 14, "G": 18, "H": 48, "I": 18, "J": 10, "K": 42})
link_sources(province)

main = wb.create_sheet("项目主表")
main_headers = ["省份", "城市/地区", "项目名称", "阶段", "时间", "对象范围", "明确数量/范围", "调查面积(km²)", "监测井数量", "采样数量", "检测指标", "污染/风险/成果结论", "项目状态", "来源ID", "可信度"]
add_sheet_table(main, [main_headers] + project_rows, "tbl_project_main")
style_sheet(main, {"A": 10, "B": 18, "C": 40, "D": 16, "E": 16, "F": 32, "G": 34, "H": 14, "I": 14, "J": 14, "K": 14, "L": 44, "M": 16, "N": 16, "O": 10})
link_sources(main)

quant = wb.create_sheet("省份量化对比")
quant_headers = ["地区/项目", "园区/对象数", "调查面积(km²)", "监测井数量", "采样数量", "检测指标", "预算/合同额(万元)", "来源ID", "说明"]
quant_rows = []
for r in province_rows:
    if any(num_or_none(x) is not None for x in [r[2], r[3], r[4], r[5], r[6]]):
        quant_rows.append([r[0], r[2], r[3], r[4], r[5], None, r[6], r[8], r[7]])
add_sheet_table(quant, [quant_headers] + quant_rows, "tbl_quant_compare")
style_sheet(quant, {"A": 20, "B": 14, "C": 16, "D": 14, "E": 14, "F": 14, "G": 18, "H": 16, "I": 50})
link_sources(quant)
qchart = BarChart()
qchart.title = "分省/项目公开数量对比"
qchart.y_axis.title = "园区/对象数"
qdata = Reference(quant, min_col=2, min_row=1, max_row=min(quant.max_row, 16))
qcats = Reference(quant, min_col=1, min_row=2, max_row=min(quant.max_row, 16))
qchart.add_data(qdata, titles_from_data=True)
qchart.set_categories(qcats)
qchart.height = 7
qchart.width = 14
quant.add_chart(qchart, "K2")

proc = wb.create_sheet("采购工作量拆解")
add_sheet_table(proc, [["项目", "预算/合同额(万元)", "范围/对象", "钻探/监测井", "土壤样品", "地下水样品", "检测指标", "质控/验收要求", "来源ID"]] + procurement_rows, "tbl_procurement")
style_sheet(proc, {"A": 38, "B": 18, "C": 34, "D": 30, "E": 24, "F": 24, "G": 38, "H": 30, "I": 16})
link_sources(proc)
pchart = BarChart()
pchart.title = "采购/预算金额对比"
pchart.y_axis.title = "万元"
pdata = Reference(proc, min_col=2, min_row=1, max_row=proc.max_row)
pcats = Reference(proc, min_col=1, min_row=2, max_row=proc.max_row)
pchart.add_data(pdata, titles_from_data=True)
pchart.set_categories(pcats)
pchart.height = 7
pchart.width = 14
proc.add_chart(pchart, "K2")

risk = wb.create_sheet("风险治理动作库")
add_sheet_table(risk, [["风险/问题", "典型成因", "治理动作", "建议KPI", "来源ID"]] + risk_rows, "tbl_risk_actions")
style_sheet(risk, {"A": 24, "B": 38, "C": 44, "D": 32, "E": 18})
link_sources(risk)

methods = wb.create_sheet("方法论字段")
add_sheet_table(methods, [["方法/字段", "定义", "使用场景", "公开样例", "来源ID"]] + method_rows, "tbl_method_fields")
style_sheet(methods, {"A": 20, "B": 48, "C": 28, "D": 28, "E": 18})
link_sources(methods)

dictionary = wb.create_sheet("字段字典")
add_sheet_table(dictionary, field_rows, "tbl_field_dictionary")
style_sheet(dictionary, {"A": 22, "B": 46, "C": 52})

gap = wb.create_sheet("待补数据清单")
gap_rows = [["优先级", "数据缺口", "为什么重要", "建议检索路径", "当前处理"]]
gap_items = [
    ["P0", "399个化工园区专项整治分省清单", "决定全国治理版图和投资机会", "生态环境部专项整治通知、各省任务分解、年度预算绩效", "只记录全国总量，不外推分省"],
    ["P0", "每个园区污染因子/风险等级", "决定治理技术路线和预算规模", "调查评估报告公示、专家评审、政府采购成果公告", "公开未披露时填未披露"],
    ["P1", "监测井坐标和井深结构", "决定长期监测网络质量", "招标文件附件、验收报告、成果报告摘要", "仅记录井数"],
    ["P1", "调查评估转入修复/管控项目比例", "评估从调查到治理的转化率", "中央生态环境资金项目储备库、地方项目公示", "单列治理机会，不编造比例"],
    ["P2", "各省完整预算/合同额", "便于估算市场空间", "中国政府采购网、公共资源交易中心、财政决算", "只收录已披露金额"],
]
for item in gap_items:
    gap_rows.append(item)
for r in province_rows:
    if r[9] == "低":
        gap_rows.append(["P2", f"{r[0]}省级园区地下水调查评估公开数字", "补齐全国覆盖地图", "省生态环境厅/生态环境状况公报/政府采购关键词检索", "本轮未检索到公开量化数据"])
add_sheet_table(gap, gap_rows, "tbl_data_gap")
style_sheet(gap, {"A": 10, "B": 42, "C": 40, "D": 48, "E": 34})

src = wb.create_sheet("来源索引")
src_headers = ["来源ID", "标题", "发布日期", "来源类型", "URL", "可引用要点"]
src_rows = [[s[h] for h in src_headers] for s in sources]
add_sheet_table(src, [src_headers] + src_rows, "tbl_sources")
style_sheet(src, {"A": 12, "B": 42, "C": 14, "D": 24, "E": 56, "F": 64})
for row in range(2, src.max_row + 1):
    src.cell(row, 5).hyperlink = src.cell(row, 5).value
    src.cell(row, 5).style = "Hyperlink"

for ws in wb.worksheets:
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

wb.save(XLSX)

check = load_workbook(XLSX, data_only=False)
assert len(check.sheetnames) == 11, check.sheetnames
assert check["项目主表"].max_row == len(project_rows) + 1
assert check["省份覆盖清单"].max_row == len(province_rows) + 1
assert check["来源索引"].max_row == len(sources) + 1
assert check["管理看板"]["B11"].value == 399
assert check["项目主表"]["H6"].value == 1062.62
assert check["项目主表"]["I6"].value == 163
assert check["采购工作量拆解"]["B4"].value == 165.0
assert check["来源索引"]["E2"].hyperlink is not None
assert len(check["管理看板"]._charts) >= 1
assert len(check["省份量化对比"]._charts) >= 1
assert len(check["采购工作量拆解"]._charts) >= 1

top_quant = [r for r in province_rows if r[2] not in (None, "未披露")][:12]
md_lines = [
    f"# 全国产业园区地下水污染状况调查项目数据表（完整版）",
    "",
    f"- 生成日期：{date.today().isoformat()}",
    f"- Excel：`{XLSX}`",
    f"- 口径：产业园区/化工园区/工业园区地下水环境状况调查评估、详细调查、质控、专项整治。",
    f"- 数据原则：只采用公开可核验来源；未披露数字统一标为“未披露”，不做推算。",
    "",
    "## 核心结论",
    "",
]
md_lines += [f"- {x[3:] if x[:2].isdigit() else x}" for x in insights]
md_lines += [
    "",
    "## 工作簿结构",
    "",
    "| Sheet | 用途 |",
    "|---|---|",
]
for sheet, purpose in [
    ("管理看板", "结论先行、核心KPI、数量对比图"),
    ("全国主线", "从基础调查到专项整治的政策/项目脉络"),
    ("省份覆盖清单", "全国31省+全国口径覆盖，明确已披露和待补数据"),
    ("项目主表", "逐项目真实数据主表"),
    ("省份量化对比", "可公开量化地区/项目横向比较"),
    ("采购工作量拆解", "预算、井、样品、检测指标、质控要求"),
    ("风险治理动作库", "从调查结果转为治理动作和KPI"),
    ("方法论字段", "一园一档、建井采样、污染羽等字段解释"),
    ("字段字典", "主表字段口径"),
    ("待补数据清单", "下一轮补全全国数据的优先级清单"),
    ("来源索引", "来源ID、发布日期、类型、URL、引用要点"),
]:
    md_lines.append(f"| {sheet} | {purpose} |")

md_lines += [
    "",
    "## 重点量化数据摘录",
    "",
    "| 地区/项目 | 园区/对象数 | 面积(km²) | 监测井 | 采样 | 预算/合同额(万元) | 来源 |",
    "|---|---:|---:|---:|---:|---:|---|",
]
for r in top_quant:
    md_lines.append(
        f"| {r[0]} | {r[2] if r[2] is not None else '未披露'} | {r[3] if r[3] is not None else '未披露'} | {r[4] if r[4] is not None else '未披露'} | {r[5] if r[5] is not None else '未披露'} | {r[6] if r[6] is not None else '未披露'} | {r[8]} |"
    )

md_lines += [
    "",
    "## 采购/工作量样本",
    "",
    "| 项目 | 预算/合同额(万元) | 工作量要点 | 来源 |",
    "|---|---:|---|---|",
]
for r in procurement_rows:
    md_lines.append(f"| {r[0]} | {r[1]} | {r[3]}；{r[4]}；{r[5]} | {r[8]} |")

md_lines += [
    "",
    "## 来源索引（节选）",
    "",
    "| 来源ID | 标题 | 日期 | URL |",
    "|---|---|---|---|",
]
for s in sources[:12]:
    md_lines.append(f"| {s['来源ID']} | {s['标题']} | {s['发布日期']} | {s['URL']} |")

md_lines += [
    "",
    "## 待补数据优先级",
    "",
    "- P0：399个专项整治园区分省清单、园区污染因子/风险等级。",
    "- P1：监测井坐标/井深结构、调查成果向治理项目转化率。",
    "- P2：各省预算/合同额和未公开省份的项目清单。",
]
MD.write_text("\n".join(md_lines), encoding="utf-8")

print(f"created: {XLSX}")
print(f"created: {MD}")
print(f"sheets: {', '.join(check.sheetnames)}")
print(f"project_rows: {len(project_rows)}; province_rows: {len(province_rows)}; sources: {len(sources)}")
